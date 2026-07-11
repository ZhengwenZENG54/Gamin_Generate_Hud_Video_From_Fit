import sys
import os
from datetime import datetime, timedelta
from fitparse import FitFile
import importlib.util
import glob
import json
import time
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ==================== 用户可修改的配置 ====================
# FIT文件路径，可以设置为具体路径或None
# 如果设置为None，将自动查找脚本所在目录下最新的.fit文件
FIT_PATH = None
# FIT_PATH = r"Gamin_Generate_Hud_Video_From_Fit\2026-04-18-18-53-11.fit"

# 原模块路径
MODULE_PATH_ALPHA = r"Gamin_Generate_Hud_Video_From_Fit\Alpha_hud_map_elevation.py"
MODULE_PATH_BETA = r"Gamin_Generate_Hud_Video_From_Fit\Beta_time_distance_elevation.py" 

# 日志文件路径
LOG_FILE = r"Gamin_Generate_Hud_Video_From_Fit\fit_video_audit_log.txt"

# Excel统计文件路径
EXCEL_FILE = r"Gamin_Generate_Hud_Video_From_Fit\fit_video_audit_log.xlsx"

# 注意：这里改为默认不生成任何视频，让用户选择
GENERATE_HUD = False    # 是否生成HUD视频
GENERATE_MAP = False    # 是否生成地图视频
GENERATE_ELEVATION = False  # 是否生成海拔视频

# 新增：视频帧率配置（默认值，可以根据需要修改）
ALPHA_HUD_FPS = 30            # HUD视频帧率
ALPHA_MAP_FPS = 5             # 地图视频帧率
ALPHA_ELEVATION_FPS = 5       # 海拔视频帧率

# Beta模块默认帧率
BETA_TIME_FPS = 1      # 时间视频帧率
BETA_DISTANCE_FPS = 5  # 距离视频帧率
BETA_ELEVATION_FPS = 5 # 海拔视频帧率
# =====================================================

def load_module_from_path(module_name, file_path):
    """从指定路径动态加载模块"""
    spec = importlib.util.spec_from_file_location(module_name, file_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module

def get_all_laps(fit_path):
    """获取fit文件中的所有lap信息"""
    print(f"正在读取FIT文件: {fit_path}")
    fit = FitFile(fit_path)
    
    laps = []
    print("\n=== FIT文件中的Lap信息 ===")
    
    for i, lap in enumerate(fit.get_messages("lap")):
        vals = lap.get_values()
        start_time = vals.get("start_time")
        elapsed = vals.get("total_elapsed_time")
        trigger = vals.get("lap_trigger")
        total_distance = vals.get("total_distance", 0.0)  # 获取总距离
        
        if start_time is not None and elapsed is not None:
            end_time = start_time + timedelta(seconds=elapsed)
            end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
            
            lap_info = {
                "index": i + 1,
                "start_time": start_time,
                "end_time": end_time,
                "elapsed_seconds": elapsed,
                "total_distance": total_distance,
                "trigger": trigger
            }
            laps.append(lap_info)
            
            print(f"[Lap {i+1}]")
            print(f"  开始时间: {start_time}")
            print(f"  结束时间: {end_time_str}")
            print(f"  持续时间: {elapsed:.1f}秒")
            print(f"  总距离: {total_distance:.2f}m ({total_distance/1000:.2f}km)")
            print(f"  触发类型: {trigger}")
            print(f"  {'-'*40}")
    
    if not laps:
        print("未找到Lap信息")
    
    return laps

def select_laps_for_generation(laps):
    """让用户选择要生成的lap"""
    if not laps:
        print("没有可用的Lap信息")
        return None
    
    print("\n请选择要生成视频的Lap（输入序号，多个用逗号分隔，q退出）:")
    for lap in laps:
        print(f"{lap['index']}. Lap {lap['index']} ({lap['elapsed_seconds']:.1f}秒, {lap['total_distance']/1000:.2f}km)")
    
    while True:
        choice = input("请输入选择: ").strip()
        
        if choice.lower() == 'q':
            return None
        
        if choice:
            try:
                selected_indices = [int(idx.strip()) for idx in choice.split(',')]
                selected_laps = []
                selected_start = None
                selected_end = None
                
                for idx in selected_indices:
                    if 1 <= idx <= len(laps):
                        lap = laps[idx-1]
                        selected_laps.append(lap)
                        
                        if selected_start is None or lap['start_time'] < selected_start:
                            selected_start = lap['start_time']
                        if selected_end is None or lap['end_time'] > selected_end:
                            selected_end = lap['end_time']
                    else:
                        print(f"警告: Lap {idx} 不存在，跳过")
                
                if selected_laps:
                    print(f"\n已选择 {len(selected_laps)} 个Lap:")
                    for lap in selected_laps:
                        print(f"  Lap {lap['index']}: {lap['start_time']} 到 {lap['end_time']}")
                    print(f"合并时间范围: {selected_start} 到 {selected_end}")
                    return selected_start, selected_end, selected_laps
                else:
                    print("没有有效的选择")
            except ValueError:
                print("输入无效，请重新输入")

def find_latest_fit_file():
    """查找当前目录下最新的.fit文件"""
    fit_files = glob.glob("*.fit")
    
    if not fit_files:
        # 如果没有找到，尝试在子目录中查找
        fit_files = glob.glob("**/*.fit", recursive=True)
    
    if not fit_files:
        return None
    
    # 按修改时间排序，获取最新的文件
    fit_files.sort(key=os.path.getmtime, reverse=True)
    latest_fit = os.path.abspath(fit_files[0])
    print(f"找到最新的FIT文件: {latest_fit}")
    return latest_fit

def get_fit_path():
    """获取FIT文件路径，如果未指定则自动查找最新文件"""
    if FIT_PATH is None:
        print("未指定FIT_PATH，正在查找当前目录下最新的.fit文件...")
        fit_path = find_latest_fit_file()
        if fit_path is None:
            print("❌ 未找到任何.fit文件")
            print("请在代码开头设置FIT_PATH，或在当前目录下放置.fit文件")
            return None
        return fit_path
    else:
        if not os.path.exists(FIT_PATH):
            print(f"❌ 找不到指定的FIT文件: {FIT_PATH}")
            return None
        return os.path.abspath(FIT_PATH)

def get_alpha_video_selection():
    """让用户选择生成哪种alpha视频"""
    print("\n请选择要生成的Alpha视频类型:")
    print("1. 只生成HUD视频（数据叠加层）")
    print("2. 只生成地图视频（轨迹动画）")
    print("3. 只生成海拔视频（爬升动画）")
    print("4. 生成HUD + 地图")
    print("5. 生成HUD + 海拔")
    print("6. 生成地图 + 海拔")
    print("7. 全部生成（HUD + 地图 + 海拔）")
    print("8. 跳过Alpha模块，直接进入Beta模块")
    print("q. 退出")
    
    while True:
        choice = input("请输入选择 (1/2/3/4/5/6/7/8/q): ").strip().lower()
        
        if choice == 'q':
            return None, None, None, True  # 用户退出
        
        if choice in ['1', '2', '3', '4', '5', '6', '7', '8']:
            if choice == '1':
                return True, False, False, False
            elif choice == '2':
                return False, True, False, False
            elif choice == '3':
                return False, False, True, False
            elif choice == '4':
                return True, True, False, False
            elif choice == '5':
                return True, False, True, False
            elif choice == '6':
                return False, True, True, False
            elif choice == '7':
                return True, True, True, False
            else:  # choice == '8'
                return False, False, False, True  # 跳过Alpha
        else:
            print("输入无效，请重新输入")

def get_beta_video_selection():
    """让用户选择生成哪种beta视频"""
    print("\n请选择要生成的Beta视频类型:")
    print("1. 只生成时间视频")
    print("2. 只生成距离视频")
    print("3. 只生成海拔视频")
    print("4. 生成时间 + 距离")
    print("5. 生成时间 + 海拔")
    print("6. 生成距离 + 海拔")
    print("7. 全部生成（时间 + 距离 + 海拔）")
    print("8. 跳过Beta模块")
    print("q. 退出")
    
    while True:
        choice = input("请输入选择 (1/2/3/4/5/6/7/8/q): ").strip().lower()
        
        if choice == 'q':
            return None, None, None, True  # 用户退出
        
        if choice in ['1', '2', '3', '4', '5', '6', '7', '8']:
            if choice == '1':
                return True, False, False, False
            elif choice == '2':
                return False, True, False, False
            elif choice == '3':
                return False, False, True, False
            elif choice == '4':
                return True, True, False, False
            elif choice == '5':
                return True, False, True, False
            elif choice == '6':
                return False, True, True, False
            elif choice == '7':
                return True, True, True, False
            else:  # choice == '8'
                return False, False, False, True  # 跳过Beta
        else:
            print("输入无效，请重新输入")

def seconds_to_hms(seconds):
    """将秒转换为时-分-秒格式"""
    if seconds is None:
        return "N/A"
    
    hours = int(seconds // 3600)
    minutes = int((seconds % 3600) // 60)
    secs = int(seconds % 60)
    
    if hours > 0:
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    else:
        return f"{minutes:02d}:{secs:02d}"

def auto_adjust_column_width(worksheet):
    """自动调整Excel列宽为最合适状态"""
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    # 计算单元格内容的长度
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # 设置列宽，留有一些边距
        adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
        worksheet.column_dimensions[column_letter].width = adjusted_width

def write_audit_excel(audit_info):
    """将审计信息写入Excel文件"""
    try:
        # 提取需要的数据
        timestamp_str = audit_info.get('timestamp', '')
        fit_file = audit_info.get('fit_file', '')
        
        # 计算连续视频总时长（秒）
        continuous_duration_sec = audit_info.get('continuous_duration_sec', 0.0)
        
        # Alpha模块数据
        generate_hud_alpha = audit_info.get('generate_hud_alpha', False)
        generate_map_alpha = audit_info.get('generate_map_alpha', False)
        generate_elev_alpha = audit_info.get('generate_elevation_alpha', False)
        
        hud_fps_alpha = audit_info.get('hud_fps_alpha') if generate_hud_alpha else ''
        map_fps_alpha = audit_info.get('map_fps_alpha') if generate_map_alpha else ''
        elev_fps_alpha = audit_info.get('elevation_fps_alpha') if generate_elev_alpha else ''
        alpha_time = audit_info.get('total_time_alpha', 0.0)
        
        # Beta模块数据
        success_beta = audit_info.get('success_beta', False)
        if success_beta:
            generate_time_beta = audit_info.get('generate_time_beta', False)
            generate_dist_beta = audit_info.get('generate_distance_beta', False)
            generate_elev_beta = audit_info.get('generate_elevation_beta', False)
            
            time_fps_beta = audit_info.get('time_fps_beta') if generate_time_beta else ''
            dist_fps_beta = audit_info.get('distance_fps_beta') if generate_dist_beta else ''
            elev_fps_beta = audit_info.get('elevation_fps_beta') if generate_elev_beta else ''
            beta_time = audit_info.get('total_time_beta', 0.0)
        else:
            # Beta模块未运行或运行失败
            time_fps_beta = ''
            dist_fps_beta = ''
            elev_fps_beta = ''
            beta_time = ''
        
        # 总耗时
        total_time = audit_info.get('total_time', 0.0)
        
        # 创建一行数据
        row_data = {
            '调用时间': timestamp_str,  # 保持和txt一致的格式
            '调用文件': fit_file,
            'Lap总时长': continuous_duration_sec,
            'Alpha_HUD 帧率': hud_fps_alpha,
            'Alpha_map帧率': map_fps_alpha,
            'Alpha_elev帧率': elev_fps_alpha,
            'Alpha耗时': alpha_time,
            'Beta_time帧率': time_fps_beta,
            'Beta_dist帧率': dist_fps_beta,
            'Beta_elev帧率': elev_fps_beta,
            'Beta耗时': beta_time,
            '总耗时': total_time
        }
        
        # 检查Excel文件是否存在
        if os.path.exists(EXCEL_FILE):
            try:
                # 尝试读取现有文件
                df = pd.read_excel(EXCEL_FILE)
                
                # 检查列名是否一致
                expected_columns = [
                    '调用时间', '调用文件', 'Lap总时长', 
                    'Alpha_HUD 帧率', 'Alpha_map帧率', 'Alpha_elev帧率', 'Alpha耗时',
                    'Beta_time帧率', 'Beta_dist帧率', 'Beta_elev帧率', 'Beta耗时', 
                    '总耗时'
                ]
                
                # 如果列不一致，重新创建
                if list(df.columns) != expected_columns:
                    print(f"警告: Excel文件列格式不一致，重新创建文件")
                    df = pd.DataFrame(columns=expected_columns)
            except Exception as e:
                print(f"警告: 读取Excel文件失败，重新创建: {e}")
                df = pd.DataFrame(columns=[
                    '调用时间', '调用文件', 'Lap总时长', 
                    'Alpha_HUD 帧率', 'Alpha_map帧率', 'Alpha_elev帧率', 'Alpha耗时',
                    'Beta_time帧率', 'Beta_dist帧率', 'Beta_elev帧率', 'Beta耗时', 
                    '总耗时'
                ])
        else:
            # 创建新的DataFrame
            df = pd.DataFrame(columns=[
                '调用时间', '调用文件', 'Lap总时长', 
                'Alpha_HUD 帧率', 'Alpha_map帧率', 'Alpha_elev帧率', 'Alpha耗时',
                'Beta_time帧率', 'Beta_dist帧率', 'Beta_elev帧率', 'Beta耗时', 
                '总耗时'
            ])
        
        # 添加新行
        df = pd.concat([df, pd.DataFrame([row_data])], ignore_index=True)
        
        # 保存到Excel
        df.to_excel(EXCEL_FILE, index=False)
        
        # 使用openpyxl调整列宽
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        auto_adjust_column_width(ws)
        wb.save(EXCEL_FILE)
        
        print(f"✅ 审计数据已保存到Excel: {os.path.abspath(EXCEL_FILE)}")
        
    except Exception as e:
        print(f"❌ 写入Excel文件失败: {e}")
        print(traceback.format_exc())

def write_audit_log(audit_info):
    """将审计信息写入日志文件"""
    try:
        # 检查日志文件是否存在，不存在则创建
        file_exists = os.path.exists(LOG_FILE)
        
        with open(LOG_FILE, 'a', encoding='utf-8') as f:
            if not file_exists:
                # 写入表头
                f.write("="*80 + "\n")
                f.write("FIT视频生成审计日志\n")
                f.write("="*80 + "\n\n")
            
            # 写入本次调用信息
            f.write(f"[调用时间] {audit_info['timestamp']}\n")
            f.write(f"FIT文件: {audit_info['fit_file']}\n")
            
            if audit_info['selected_laps']:
                f.write(f"选择的Lap: {', '.join(map(str, audit_info['selected_lap_indices']))}\n")
                f.write(f"选择的Lap详细信息:\n")
                
                # 对选中的Lap按开始时间排序
                sorted_laps = sorted(audit_info['selected_laps'], key=lambda x: x['start_time'])
                
                for lap in sorted_laps:
                    lap_seconds = lap.get('elapsed_seconds', 0)
                    hms = seconds_to_hms(lap_seconds)
                    f.write(f"  Lap {lap.get('index', 'N/A')}: {hms} ({lap_seconds:.1f}秒)\n")
                
                # 计算连续视频的总时长
                if len(sorted_laps) > 0:
                    # 获取最早的开始时间和最晚的结束时间
                    first_start = min(lap['start_time'] for lap in sorted_laps)
                    last_end = max(lap['end_time'] for lap in sorted_laps)
                    continuous_duration = (last_end - first_start).total_seconds()
                    
                    # 计算选择的Lap总时长（简单相加）
                    selected_laps_total = sum(lap.get('elapsed_seconds', 0) for lap in sorted_laps)
                    
                    # 计算中间间隔的时长
                    gap_duration = continuous_duration - selected_laps_total
                    
                    continuous_hms = seconds_to_hms(continuous_duration)
                    selected_hms = seconds_to_hms(selected_laps_total)
                    gap_hms = seconds_to_hms(gap_duration)
                    
                    f.write(f"连续视频时间范围: {first_start} 到 {last_end}\n")
                    f.write(f"连续视频总时长: {continuous_hms} ({continuous_duration:.1f}秒)\n")
                    f.write(f"选中Lap总时长: {selected_hms} ({selected_laps_total:.1f}秒)\n")
                    f.write(f"中间间隔时长: {gap_hms} ({gap_duration:.1f}秒)\n")
                    
                    # 将连续视频总时长保存到audit_info中，用于Excel写入
                    audit_info['continuous_duration_sec'] = continuous_duration
            else:
                f.write("选择的Lap: 无\n")
            
            # 模块alpha生成结果
            f.write("\n=== 模块alpha生成结果 ===\n")
            video_types_alpha = []
            if audit_info.get('generate_hud_alpha'):
                video_types_alpha.append("HUD视频")
            if audit_info.get('generate_map_alpha'):
                video_types_alpha.append("地图视频")
            if audit_info.get('generate_elevation_alpha'):
                video_types_alpha.append("海拔视频")
            
            f.write(f"生成视频类型: {' + '.join(video_types_alpha) if video_types_alpha else '无'}\n")
            
            # 视频帧率设置
            f.write("视频帧率设置:\n")
            if audit_info.get('generate_hud_alpha'):
                f.write(f"  HUD视频: {audit_info.get('hud_fps_alpha', 'N/A')} FPS\n")
            if audit_info.get('generate_map_alpha'):
                f.write(f"  地图视频: {audit_info.get('map_fps_alpha', 'N/A')} FPS\n")
            if audit_info.get('generate_elevation_alpha'):
                f.write(f"  海拔视频: {audit_info.get('elevation_fps_alpha', 'N/A')} FPS\n")
            
            # 时间统计
            f.write(f"时间统计:\n")
            if audit_info.get('hud_time_a') is not None and audit_info.get('generate_hud_alpha'):
                f.write(f"  HUD视频生成时间: {audit_info['hud_time_a']:.2f}秒\n")
            if audit_info.get('map_time_a') is not None and audit_info.get('generate_map_alpha'):
                f.write(f"  地图视频生成时间: {audit_info['map_time_a']:.2f}秒\n")
            if audit_info.get('elevation_time_a') is not None and audit_info.get('generate_elevation_alpha'):
                f.write(f"  海拔视频生成时间: {audit_info['elevation_time_a']:.2f}秒\n")
            
            f.write(f"  模块alpha总耗时: {audit_info.get('total_time_alpha', 0):.2f}秒\n")
            
            # 生成的帧数
            f.write("生成的帧数:\n")
            if audit_info.get('hud_frame_count_alpha') and audit_info.get('generate_hud_alpha'):
                f.write(f"  HUD视频: {audit_info['hud_frame_count_alpha']} 帧\n")
            if audit_info.get('map_frame_count_alpha') and audit_info.get('generate_map_alpha'):
                f.write(f"  地图视频: {audit_info['map_frame_count_alpha']} 帧\n")
            if audit_info.get('elevation_frame_count_alpha') and audit_info.get('generate_elevation_alpha'):
                f.write(f"  海拔视频: {audit_info['elevation_frame_count_alpha']} 帧\n")
            
            # 模块alpha生成的文件路径
            f.write(f"模块alpha生成的文件:\n")
            for key, value in audit_info.get('result_alpha', {}).items():
                if isinstance(value, str) and os.path.exists(value):
                    f.write(f"  {key}: {value}\n")
            
            # 模块beta生成结果
            f.write("\n=== 模块beta生成结果 ===\n")
            video_types_beta = []
            if audit_info.get('generate_time_beta'):
                video_types_beta.append("时间视频")
            if audit_info.get('generate_distance_beta'):
                video_types_beta.append("距离视频")
            if audit_info.get('generate_elevation_beta'):
                video_types_beta.append("海拔视频")
            
            f.write(f"生成视频类型: {' + '.join(video_types_beta) if video_types_beta else '无'}\n")
            
            # 视频帧率设置
            f.write("视频帧率设置:\n")
            if audit_info.get('generate_time_beta'):
                f.write(f"  时间视频: {audit_info.get('time_fps_beta', 'N/A')} FPS\n")
            if audit_info.get('generate_distance_beta'):
                f.write(f"  距离视频: {audit_info.get('distance_fps_beta', 'N/A')} FPS\n")
            if audit_info.get('generate_elevation_beta'):
                f.write(f"  海拔视频: {audit_info.get('elevation_fps_beta', 'N/A')} FPS\n")
            
            # 时间统计
            f.write(f"时间统计:\n")
            if audit_info.get('time_time_beta') is not None and audit_info.get('generate_time_beta'):
                f.write(f"  时间视频生成时间: {audit_info['time_time_beta']:.2f}秒\n")
            if audit_info.get('distance_time_beta') is not None and audit_info.get('generate_distance_beta'):
                f.write(f"  距离视频生成时间: {audit_info['distance_time_beta']:.2f}秒\n")
            if audit_info.get('elevation_time_beta') is not None and audit_info.get('generate_elevation_beta'):
                f.write(f"  海拔视频生成时间: {audit_info['elevation_time_beta']:.2f}秒\n")
            
            f.write(f"  模块beta总耗时: {audit_info.get('total_time_beta', 0):.2f}秒\n")
            
            # 生成的帧数
            f.write("生成的帧数:\n")
            if audit_info.get('time_frame_count_beta') and audit_info.get('generate_time_beta'):
                f.write(f"  时间视频: {audit_info['time_frame_count_beta']} 帧\n")
            if audit_info.get('distance_frame_count_beta') and audit_info.get('generate_distance_beta'):
                f.write(f"  距离视频: {audit_info['distance_frame_count_beta']} 帧\n")
            if audit_info.get('elevation_frame_count_beta') and audit_info.get('generate_elevation_beta'):
                f.write(f"  海拔视频: {audit_info['elevation_frame_count_beta']} 帧\n")
            
            # 模块beta生成的文件路径
            f.write(f"模块beta生成的文件:\n")
            for key, value in audit_info.get('result_beta', {}).items():
                if isinstance(value, str) and os.path.exists(value):
                    f.write(f"  {key}: {value}\n")
            
            # 总体统计
            f.write(f"\n=== 总体统计 ===\n")
            f.write(f"总耗时: {audit_info['total_time']:.2f}秒\n")
            
            # 状态
            status_alpha = "成功" if audit_info.get('success_alpha', True) else "失败"
            status_beta = "成功" if audit_info.get('success_beta', True) else "失败"
            f.write(f"模块alpha状态: {status_alpha}")
            if audit_info.get('error_alpha'):
                f.write(f" (错误: {audit_info['error_alpha']})")
            f.write(f"\n模块beta状态: {status_beta}")
            if audit_info.get('error_beta'):
                f.write(f" (错误: {audit_info['error_beta']})")
            
            f.write("\n")
            f.write("-"*80 + "\n\n")
        
        print(f"✅ 审计日志已保存到: {os.path.abspath(LOG_FILE)}")
        
        # 如果alpha运行成功，将数据写入Excel
        if audit_info.get('success_alpha'):
            write_audit_excel(audit_info)
        
    except Exception as e:
        print(f"❌ 写入审计日志失败: {e}")

def calculate_expected_frames(duration_sec, fps):
    """计算预期的总帧数"""
    if duration_sec is None or fps is None:
        return None
    return int(duration_sec * fps) + 1  # 加1是因为包含开始和结束帧

def get_alpha_fps_settings():
    """获取alpha模块视频帧率设置"""
    print("\n" + "="*50)
    print("模块alpha视频帧率设置（使用默认值或自定义）:")
    print("="*50)
    print(f"当前默认帧率:")
    print(f"  HUD视频: {ALPHA_HUD_FPS} FPS")
    print(f"  地图视频: {ALPHA_MAP_FPS} FPS")
    print(f"  海拔视频: {ALPHA_ELEVATION_FPS} FPS")
    print("\n1. 使用默认帧率")
    print("2. 自定义帧率")
    print("q. 退出")
    
    while True:
        choice = input("请选择 (1/2/q): ").strip().lower()
        
        if choice == 'q':
            return None, None, None
        
        if choice in ['1', '2']:
            if choice == '1':
                print(f"使用默认帧率: HUD={ALPHA_HUD_FPS}, 地图={ALPHA_MAP_FPS}, 海拔={ALPHA_ELEVATION_FPS} FPS")
                return ALPHA_HUD_FPS, ALPHA_MAP_FPS, ALPHA_ELEVATION_FPS
            else:
                # 自定义帧率
                try:
                    print("\n请输入各视频的帧率（正整数）:")
                    hud_fps_input = input(f"HUD视频帧率 (默认{ALPHA_HUD_FPS}): ").strip()
                    map_fps_input = input(f"地图视频帧率 (默认{ALPHA_MAP_FPS}): ").strip()
                    elevation_fps_input = input(f"海拔视频帧率 (默认{ALPHA_ELEVATION_FPS}): ").strip()
                    
                    # 使用默认值或自定义值
                    hud_fps = int(hud_fps_input) if hud_fps_input else ALPHA_HUD_FPS
                    map_fps = int(map_fps_input) if map_fps_input else ALPHA_MAP_FPS
                    elevation_fps = int(elevation_fps_input) if elevation_fps_input else ALPHA_ELEVATION_FPS
                    
                    # 验证帧率有效性
                    if hud_fps <= 0 or map_fps <= 0 or elevation_fps <= 0:
                        print("错误: 帧率必须是正整数")
                        continue
                    
                    print(f"自定义帧率设置: HUD={hud_fps}, 地图={map_fps}, 海拔={elevation_fps} FPS")
                    return hud_fps, map_fps, elevation_fps
                except ValueError:
                    print("错误: 请输入有效的数字")
        else:
            print("输入无效，请重新输入")

def get_beta_fps_settings():
    """获取beta模块视频帧率设置"""
    print("\n" + "="*50)
    print("模块beta视频帧率设置（使用默认值或自定义）:")
    print("="*50)
    print(f"当前默认帧率:")
    print(f"  时间视频: {BETA_TIME_FPS} FPS")
    print(f"  距离视频: {BETA_DISTANCE_FPS} FPS")
    print(f"  海拔视频: {BETA_ELEVATION_FPS} FPS")
    print("\n1. 使用默认帧率")
    print("2. 自定义帧率")
    print("q. 退出")
    
    while True:
        choice = input("请选择 (1/2/q): ").strip().lower()
        
        if choice == 'q':
            return None, None, None
        
        if choice in ['1', '2']:
            if choice == '1':
                print(f"使用默认帧率: 时间={BETA_TIME_FPS}, 距离={BETA_DISTANCE_FPS}, 海拔={BETA_ELEVATION_FPS} FPS")
                return BETA_TIME_FPS, BETA_DISTANCE_FPS, BETA_ELEVATION_FPS
            else:
                # 自定义帧率
                try:
                    print("\n请输入各视频的帧率（正整数）:")
                    time_fps_input = input(f"时间视频帧率 (默认{BETA_TIME_FPS}): ").strip()
                    distance_fps_input = input(f"距离视频帧率 (默认{BETA_DISTANCE_FPS}): ").strip()
                    elevation_fps_input = input(f"海拔视频帧率 (默认{BETA_ELEVATION_FPS}): ").strip()
                    
                    # 使用默认值或自定义值
                    time_fps = int(time_fps_input) if time_fps_input else BETA_TIME_FPS
                    distance_fps = int(distance_fps_input) if distance_fps_input else BETA_DISTANCE_FPS
                    elevation_fps = int(elevation_fps_input) if elevation_fps_input else BETA_ELEVATION_FPS
                    
                    # 验证帧率有效性
                    if time_fps <= 0 or distance_fps <= 0 or elevation_fps <= 0:
                        print("错误: 帧率必须是正整数")
                        continue
                    
                    print(f"自定义帧率设置: 时间={time_fps}, 距离={distance_fps}, 海拔={elevation_fps} FPS")
                    return time_fps, distance_fps, elevation_fps
                except ValueError:
                    print("错误: 请输入有效的数字")
        else:
            print("输入无效，请重新输入")

def run_module_alpha(fit_path, lap_start, lap_end, generate_hud, generate_map, generate_elevation, 
                 hud_fps, map_fps, elevation_fps, audit_info):
    """运行模块alpha并收集结果"""
    print("\n" + "="*60)
    print("开始运行模块alpha")
    print("="*60)
    
    try:
        # 检查模块文件是否存在
        if not os.path.exists(MODULE_PATH_ALPHA):
            error_msg = f"找不到模块alpha文件: {MODULE_PATH_ALPHA}"
            print(f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
        
        # 动态导入模块alpha
        module_alpha = load_module_from_path("module_alpha", MODULE_PATH_ALPHA)
        print("✅ 成功导入模块alpha")
        
        # 计算预期的帧数
        duration_sec = (lap_end - lap_start).total_seconds()
        hud_frame_count = calculate_expected_frames(duration_sec, hud_fps) if generate_hud else None
        map_frame_count = calculate_expected_frames(duration_sec, map_fps) if generate_map else None
        elevation_frame_count = calculate_expected_frames(duration_sec, elevation_fps) if generate_elevation else None
        
        # 更新审计信息
        audit_info.update({
            'generate_hud_alpha': generate_hud,
            'generate_map_alpha': generate_map,
            'generate_elevation_alpha': generate_elevation,
            'hud_fps_alpha': hud_fps,
            'map_fps_alpha': map_fps,
            'elevation_fps_alpha': elevation_fps,
            'hud_frame_count_alpha': hud_frame_count,
            'map_frame_count_alpha': map_frame_count,
            'elevation_frame_count_alpha': elevation_frame_count
        })
        
        # 显示开始信息
        print(f"FIT文件: {fit_path}")
        print(f"时间范围: {lap_start} 到 {lap_end}")
        print(f"时长: {duration_sec:.1f}秒")
        
        if generate_hud:
            print(f"生成HUD视频: 是 (FPS: {hud_fps}, 预期帧数: {hud_frame_count})")
        if generate_map:
            print(f"生成地图视频: 是 (FPS: {map_fps}, 预期帧数: {map_frame_count})")
        if generate_elevation:
            print(f"生成海拔视频: 是 (FPS: {elevation_fps}, 预期帧数: {elevation_frame_count})")
        
        # 记录开始时间
        start_time = time.time()
        
        # 调用模块alpha的生成函数
        result_alpha = module_alpha.generate_hud_map_elevation_video(
            fit_path=fit_path,
            lap_start=lap_start,
            lap_end=lap_end,
            generate_hud=generate_hud,
            generate_map=generate_map,
            generate_elevation=generate_elevation,
            hud_fps=hud_fps,
            map_fps=map_fps,
            elevation_fps=elevation_fps
        )
        
        # 记录结束时间
        end_time = time.time()
        total_time_alpha = end_time - start_time
        
        # 更新审计信息
        audit_info.update({
            'result_alpha': result_alpha,
            'total_time_alpha': total_time_alpha,
            'success_alpha': True
        })
        
        print(f"\n模块alpha运行完成，耗时: {total_time_alpha:.2f}秒")
        
        # 显示结果
        if result_alpha:
            print("\n模块alpha生成结果:")
            for key, value in result_alpha.items():
                print(f"  {key}: {value}")
        
        return {'success': True, 'result': result_alpha, 'total_time': total_time_alpha}
        
    except Exception as e:
        print(f"❌ 模块alpha运行失败: {e}")
        traceback.print_exc()
        
        error_msg = str(e)
        audit_info.update({
            'success_alpha': False,
            'error_alpha': error_msg
        })
        
        return {'success': False, 'error': error_msg}

def run_module_beta(fit_path, lap_start, lap_end, audit_info, generate_time, generate_distance, generate_elevation, 
                   time_fps=None, distance_fps=None, elevation_fps=None):
    """运行模块beta并收集结果"""
    print("\n" + "="*60)
    print("开始运行模块beta")
    print("="*60)
    
    try:
        # 检查模块文件是否存在
        if not os.path.exists(MODULE_PATH_BETA):
            error_msg = f"找不到模块beta文件: {MODULE_PATH_BETA}"
            print(f"❌ {error_msg}")
            return {'success': False, 'error': error_msg}
        
        # 动态导入模块beta
        module_beta = load_module_from_path("module_beta", MODULE_PATH_BETA)
        print("✅ 成功导入模块beta")
        
        # 如果传入了自定义帧率，设置到模块中
        # 注意：这里我们直接将帧率设置到模块的属性中，这样模块内部就能使用这些帧率
        if time_fps is not None:
            # 检查模块是否有FPS_TIME属性，如果有就设置
            if hasattr(module_beta, 'FPS_TIME'):
                module_beta.FPS_TIME = time_fps
                print(f"已设置时间视频帧率: {time_fps} FPS")
        
        if distance_fps is not None:
            if hasattr(module_beta, 'FPS_DISTANCE'):
                module_beta.FPS_DISTANCE = distance_fps
                print(f"已设置距离视频帧率: {distance_fps} FPS")
        
        if elevation_fps is not None:
            if hasattr(module_beta, 'FPS_ELEVATION'):
                module_beta.FPS_ELEVATION = elevation_fps
                print(f"已设置海拔视频帧率: {elevation_fps} FPS")
        
        # 如果用户没有选择生成某个视频，我们将其帧率设为None，这样在审计信息中就不会记录
        # 但我们需要记录用户实际选择的帧率设置
        audit_time_fps = time_fps if generate_time else ''
        audit_distance_fps = distance_fps if generate_distance else ''
        audit_elevation_fps = elevation_fps if generate_elevation else ''
        
        # 计算预期的帧数
        duration_sec = (lap_end - lap_start).total_seconds()
        time_frame_count = calculate_expected_frames(duration_sec, time_fps) if generate_time and time_fps is not None else None
        distance_frame_count = calculate_expected_frames(duration_sec, distance_fps) if generate_distance and distance_fps is not None else None
        elevation_frame_count = calculate_expected_frames(duration_sec, elevation_fps) if generate_elevation and elevation_fps is not None else None
        
        # 更新审计信息
        audit_info.update({
            'generate_time_beta': generate_time,
            'generate_distance_beta': generate_distance,
            'generate_elevation_beta': generate_elevation,
            'time_fps_beta': audit_time_fps,
            'distance_fps_beta': audit_distance_fps,
            'elevation_fps_beta': audit_elevation_fps,
            'time_frame_count_beta': time_frame_count,
            'distance_frame_count_beta': distance_frame_count,
            'elevation_frame_count_beta': elevation_frame_count
        })
        
        # 显示开始信息
        print(f"FIT文件: {fit_path}")
        print(f"时间范围: {lap_start} 到 {lap_end}")
        print(f"时长: {duration_sec:.1f}秒")
        
        if generate_time:
            # 显示实际的帧率设置
            actual_time_fps = time_fps if time_fps is not None else (getattr(module_beta, 'FPS_TIME', 1) if hasattr(module_beta, 'FPS_TIME') else 1)
            print(f"生成时间视频: 是 (FPS: {actual_time_fps}, 预期帧数: {time_frame_count})")
        if generate_distance:
            actual_distance_fps = distance_fps if distance_fps is not None else (getattr(module_beta, 'FPS_DISTANCE', 5) if hasattr(module_beta, 'FPS_DISTANCE') else 5)
            print(f"生成距离视频: 是 (FPS: {actual_distance_fps}, 预期帧数: {distance_frame_count})")
        if generate_elevation:
            actual_elevation_fps = elevation_fps if elevation_fps is not None else (getattr(module_beta, 'FPS_ELEVATION', 5) if hasattr(module_beta, 'FPS_ELEVATION') else 5)
            print(f"生成海拔视频: 是 (FPS: {actual_elevation_fps}, 预期帧数: {elevation_frame_count})")
        
        # 记录开始时间
        start_time = time.time()
        
        # 尝试调用模块beta的生成函数
        result_beta = None
        
        # 方法1: 尝试调用generate_videos_from_fit函数
        if hasattr(module_beta, 'generate_videos_from_fit'):
            print("调用模块beta的generate_videos_from_fit函数...")
            result_beta = module_beta.generate_videos_from_fit(
                fit_path=fit_path,
                lap_start=lap_start,
                lap_end=lap_end,
                generate_time=generate_time,
                generate_distance=generate_distance,
                generate_elevation=generate_elevation
            )
        # 方法2: 尝试调用main函数
        elif hasattr(module_beta, 'main'):
            print("警告: 模块beta没有generate_videos_from_fit函数，尝试调用main函数...")
            # 由于main函数可能需要用户交互，这里可能需要特殊处理
            result_beta = module_beta.main()
        else:
            error_msg = "模块beta没有可调用的生成函数"
            print(f"❌ {error_msg}")
            audit_info.update({
                'success_beta': False,
                'error_beta': error_msg
            })
            return {'success': False, 'error': error_msg}
        
        # 记录结束时间
        end_time = time.time()
        total_time_beta = end_time - start_time
        
        # 假设 beta 模块在失败时会返回一个带有 'error' 键的字典
        if isinstance(result_beta, dict) and result_beta.get('error'):
            raise RuntimeError(f"Beta模块返回错误: {result_beta['error']}")
        elif result_beta is None:
            raise RuntimeError("Beta模块返回空值，可能未正确执行")
        
        # 更新审计信息
        audit_info.update({
            'result_beta': result_beta or {},
            'total_time_beta': total_time_beta,
            'success_beta': True
        })
        
        print(f"\n模块beta运行完成，耗时: {total_time_beta:.2f}秒")
        
        # 显示结果
        if result_beta and isinstance(result_beta, dict):
            print("\n模块beta生成结果:")
            for key, value in result_beta.items():
                print(f"  {key}: {value}")
        
        return {'success': True, 'result': result_beta, 'total_time': total_time_beta}
        
    except Exception as e:
        print(f"❌ 模块beta运行失败: {e}")
        traceback.print_exc()
        
        error_msg = str(e)
        audit_info.update({
            'success_beta': False,
            'error_beta': error_msg
        })
        
        return {'success': False, 'error': error_msg}

def main():
    """主函数"""
    # 记录总开始时间
    total_start_time = time.time()
    
    # 初始化审计信息
    audit_info = {
        'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        'fit_file': None,
        'selected_lap_indices': [],
        'selected_laps': [],
        'total_time': 0,
        'success_alpha': True,
        'success_beta': True,
        'error_alpha': None,
        'error_beta': None
    }
    
    try:
        # 获取FIT文件路径
        fit_path = get_fit_path()
        if fit_path is None:
            error_msg = "无法获取FIT文件路径"
            print(f"❌ {error_msg}")
            audit_info['total_time'] = time.time() - total_start_time
            write_audit_log(audit_info)
            return
        
        audit_info['fit_file'] = os.path.basename(fit_path)
        
        # 获取所有lap信息
        laps = get_all_laps(fit_path)
        
        if not laps:
            error_msg = "文件中没有Lap信息，无法生成视频"
            print(error_msg)
            audit_info['total_time'] = time.time() - total_start_time
            write_audit_log(audit_info)
            return
        else:
            # 让用户选择要生成的lap
            selection = select_laps_for_generation(laps)
            if selection is None:
                print("用户取消操作")
                audit_info['total_time'] = time.time() - total_start_time
                write_audit_log(audit_info)
                return
            
            lap_start, lap_end, selected_laps = selection
            audit_info['selected_laps'] = selected_laps
            audit_info['selected_lap_indices'] = [lap['index'] for lap in selected_laps]
        
        # 步骤1: 让用户选择要生成的Alpha视频
        generate_hud, generate_map, generate_elevation, skip_alpha = get_alpha_video_selection()
        if generate_hud is None:  # 用户选择退出
            print("用户取消操作")
            audit_info['total_time'] = time.time() - total_start_time
            write_audit_log(audit_info)
            return
        
        # 检查是否跳过Alpha
        if skip_alpha:
            print("跳过Alpha模块")
            audit_info.update({
                'generate_hud_alpha': False,
                'generate_map_alpha': False,
                'generate_elevation_alpha': False,
                'success_alpha': True  # 标记为成功（用户选择跳过）
            })
        else:
            # 步骤2: 让用户决定alpha视频帧率
            hud_fps, map_fps, elevation_fps = get_alpha_fps_settings()
            if hud_fps is None:  # 用户选择退出
                print("用户取消操作")
                audit_info['total_time'] = time.time() - total_start_time
                write_audit_log(audit_info)
                return
        
        # 步骤3: 让用户选择要生成的beta视频
        generate_time, generate_distance, generate_elevation_beta, skip_beta = get_beta_video_selection()
        if generate_time is None:  # 用户选择退出
            print("用户取消操作")
            audit_info['total_time'] = time.time() - total_start_time
            write_audit_log(audit_info)
            return
        
        # 检查是否跳过Beta
        if skip_beta:
            print("跳过Beta模块")
            audit_info.update({
                'generate_time_beta': False,
                'generate_distance_beta': False,
                'generate_elevation_beta': False,
                'success_beta': True  # 标记为成功（用户选择跳过）
            })
        else:
            # 步骤4: 让用户决定beta视频帧率
            beta_time_fps, beta_distance_fps, beta_elevation_fps = get_beta_fps_settings()
            if beta_time_fps is None:  # 用户选择退出
                print("用户取消操作")
                audit_info['total_time'] = time.time() - total_start_time
                write_audit_log(audit_info)
                return
        
        # 检查是否两个模块都跳过
        if skip_alpha and skip_beta:
            print("两个模块都跳过，程序结束")
            audit_info['total_time'] = time.time() - total_start_time
            write_audit_log(audit_info)
            return
        
        # 运行模块alpha（如果未跳过）
        result_alpha = None
        if not skip_alpha:
            result_alpha = run_module_alpha(fit_path, lap_start, lap_end, generate_hud, generate_map, generate_elevation,
                                  hud_fps, map_fps, elevation_fps, audit_info)
        else:
            # 如果跳过alpha，但需要运行beta，我们需要设置一些基本信息
            audit_info.update({
                'success_alpha': True,  # 标记为成功（虽然未运行，但这是用户的选择）
                'total_time_alpha': 0.0
            })
        
        # 运行模块beta（如果未跳过）
        result_beta = None
        if not skip_beta:
            result_beta = run_module_beta(fit_path, lap_start, lap_end, audit_info, 
                                        generate_time, generate_distance, generate_elevation_beta,
                                        beta_time_fps, beta_distance_fps, beta_elevation_fps)
        else:
            # 如果跳过beta，我们需要设置一些基本信息
            audit_info.update({
                'success_beta': True,  # 标记为成功（虽然未运行，但这是用户的选择）
                'total_time_beta': 0.0
            })
        
        # 计算总时间
        total_end_time = time.time()
        audit_info['total_time'] = total_end_time - total_start_time
        
        # 写入审计日志
        write_audit_log(audit_info)
        
        # 显示最终结果
        print("\n" + "="*60)
        print("处理完成")
        print("="*60)
        if not skip_alpha:
            print(f"模块alpha状态: {'成功' if result_alpha and result_alpha.get('success') else '失败'}")
        else:
            print("模块alpha状态: 已跳过")
        
        if not skip_beta:
            print(f"模块beta状态: {'成功' if result_beta and result_beta.get('success') else '失败'}")
        else:
            print("模块beta状态: 已跳过")
        
        print(f"总耗时: {audit_info['total_time']:.2f}秒")
        print(f"审计日志: {os.path.abspath(LOG_FILE)}")
        print(f"Excel统计: {os.path.abspath(EXCEL_FILE)}")
        
    except Exception as e:
        print(f"❌ 主程序发生错误: {e}")
        traceback.print_exc()
        
        audit_info['total_time'] = time.time() - total_start_time
        write_audit_log(audit_info)

if __name__ == "__main__":
    main()